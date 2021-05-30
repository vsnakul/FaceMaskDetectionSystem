import openpyxl
from flask import Flask, render_template, Response
import cv2

from flask import Flask, render_template, Response, request, url_for, current_app,flash
from flask_uploads import UploadSet, configure_uploads, IMAGES
from werkzeug.exceptions import abort, HTTPException
from werkzeug.utils import redirect

import detect_mask_video
import global_value
import pickledb
from detect_mask_video import VideoCamera
from detect_mask_image import image_detect
from facial_recognition_module import FacialIdentificationSystem
from detect_mask_video_for_implementation import VideoCamera2
from global_value import global_counter
import pyrebase
import os
import requests
import time
from datetime import date
from datetime import datetime
import PIL



app = Flask(__name__)
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'
photos = UploadSet('photos', IMAGES)
app.config['UPLOADED_PHOTOS_DEST'] = "examples"
configure_uploads(app, photos)

firebaseConfig = {
    "apiKey": "AIzaSyDlaeLQS7Ud2p4kOe_mWIT3sGtMjXy_JcQ",
    "authDomain": "facereg-813c9.firebaseapp.com",
    "databaseURL": "https://facereg-813c9-default-rtdb.firebaseio.com",
    "projectId": "facereg-813c9",
    "storageBucket": "facereg-813c9.appspot.com",
    "messagingSenderId": "62284980449",
    "appId": "1:62284980449:web:52750fc2c687643d7f57a0",
    "measurementId": "G-01P1W24GBX",
    "serviceAccount": "serviceAccountCredentials.json"
  }



@app.route('/')
def index():
    print("index() method called")
    return render_template('index.html')


def gen(detect_mask_video):
    print("gen() method called")
    while True:
        # get camera frame
        frame = detect_mask_video.get_frame()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')


def gen2(detect_mask_video_for_implementation):
    print("Inside gen2()")
    book = openpyxl.load_workbook('dummy_db.xlsx')
    sheet = book.active
    sheet['A2']=0
    book.save('dummy_db.xlsx')
    print("0 value added")
    isUploaded=False
    while True:
        jpeg, counter, cv2 = detect_mask_video_for_implementation.get_frame()
        frame=jpeg.tobytes()
        if counter >= 10:
            print("inside counter condition and counter value is ",counter)
            #cv2.destroyAllWindows()
            print("redirection started")
            cv2.destroyAllWindows()
            if(isUploaded==False):
                isUploaded=True
                firebase = pyrebase.initialize_app(firebaseConfig)
                realtimeDatabase = firebase.database()

                today = date.today()

                # dd/mm/YY
                d1 = today.strftime("%d/%m/%Y")

                now = datetime.now()

                current_time = now.strftime("%H:%M:%S")

                current_time_in_millisec=round(time.time() * 100)
                data={"date":str(d1), "time": str(current_time), "image": str(current_time_in_millisec)}
                realtimeDatabase.child("log").set(data)
                storage = firebase.storage()
                path_on_cloud = "images/" + str(current_time_in_millisec) + ".jpeg"
                storage.child(path_on_cloud).put("test.jpg")


            book = openpyxl.load_workbook('dummy_db.xlsx')

            sheet = book.active
            sheet['A2'] = 1
            book.save('dummy_db.xlsx')
            print("1 value added")

            print("redirection ended")

        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')



def generate_frames_for_identification(facial_recognition_module):
    while True:
        print("geenrater framres inside whule")
        frame = facial_recognition_module.generate_frames_to_show()

        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')


@app.route('/recogniseEngine')
def index2():
    # rendering webpage
    print("index2() method called")

    return render_template('videocapture.html')


@app.route('/recogniseImage')
def index3():
    # rendering webpage
    print("index3() method called")
    # image_detect()
    return render_template('videocapture.html')

@app.route('/next_stage')
def next_stage():
    book = openpyxl.load_workbook('dummy_db.xlsx')
    sheet = book.active
    cellValue = sheet['A2']
    cell = int(cellValue.value)
    if cell==1:
        print("inside next stage cell=1")
        return redirect(url_for('facial_recognition'))
    else:
        print("inside next stage else")
        return render_template('implementation_page.html',toast='You cannot because you dont have mask')




@app.route('/facial_recognition')
def facial_recognition():

    print("index4() method called")
    #book = openpyxl.load_workbook('dummy_db.xlsx')
    #sheet = book.active
    #cellValue = sheet['A1']
    #cell = int(cellValue.value)
    #if cell==1:
    return render_template('facial_recognition.html')
    #else:
        #print("#########")
        #flash("You cannnot perform facial recognition ,because you dont have a mask")
        #return "OK"






# @app.route('/getJavascriptData',methods=["POST"])
# def get_javascript_data():
#      actualFileName=request.json['filename']
#      print("inside javascript method")
#      print(actualFileName," javascript value is found")
#      image_detect(actualFileName)
#      return "OK";

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    print("upload() method called")
    if request.method == "POST" and 'photo' in request.files:
        filename = photos.save(request.files['photo'])
        print("Photo saved Succesfully inside the server", filename)
        image_detect(filename)
    return render_template('videocapture.html')


@app.route('/video_feed')
def video_feed():
    print("video_feed() method called")
    return Response(gen(VideoCamera()),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/facial_recognition_feed')
def facial_recognition_feed():
    print("facial_recognition_feed() method called")
    return Response(generate_frames_for_identification(FacialIdentificationSystem()),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/implementation_part')
def implementation_part():
    print("Reached Implementation Part...")
    return render_template('implementation_page.html')


@app.route('/video_feed_for_implentation_part')
def video_feed_for_implentation_part():
    print("inside implentation video feed")
    return Response(gen2(VideoCamera2()),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


if __name__ == '__main__':
    # defining server ip address and port
    app.run(debug=False)
